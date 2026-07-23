"""Backtesting API routes.

All routes are mounted under ``/api/backtesting``.
Heavy computation is offloaded via ``asyncio.to_thread`` with a
120-second timeout guard.
"""

from __future__ import annotations

import asyncio
import json
import logging
import queue
import re
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

from fastapi import APIRouter, Body, HTTPException, Query, Request
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

from src import backtesting as _bt
from src.backtesting.zip_export import (
    ExportSizeLimitExceeded,
    build_rolling_zip,
    build_single_backtest_zip,
    build_summary,
    save_rolling_backtest_zip,
)
from src.orchestrator.common.db_config import get_db2, get_db3
from src.orchestrator.common.sqlite import connect_read
from src.portfolio.currency import get_available_display_currencies
from src.portfolio.performance import get_risk_free_rate
from src.web_app.security import (
    AppSettings,
    PathPolicyError,
    configured_database_policy,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/backtesting", tags=["backtesting"])

# ---------------------------------------------------------------------------
# Concurrency guard
# ---------------------------------------------------------------------------
_MAX_CONCURRENT = 2
_semaphore = asyncio.Semaphore(_MAX_CONCURRENT)
_APP_SETTINGS = AppSettings.from_env()
_DB_PATH_POLICY = configured_database_policy(_APP_SETTINGS.allowed_data_roots)
_BACKTEST_ROOT = (
    Path(__file__).resolve().parents[2] / "data" / "Backtests"
).resolve(strict=False)
_BACKTEST_ID = re.compile(r"^\d{8}_\d{6}(?:_[0-9a-f]{8})?$")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _resolve_db(db_path: str = "") -> str:
    """Resolve a database path, falling back to the configured DB2."""
    supplied = db_path.strip()
    resolved = get_db2() if supplied in {"", "default", "standardized"} else supplied
    if not resolved:
        raise HTTPException(status_code=503, detail="No database configured.")
    try:
        return str(_DB_PATH_POLICY.authorize_database(resolved))
    except PathPolicyError as exc:
        status = 400 if supplied not in {"", "default", "standardized"} else 503
        raise HTTPException(status_code=status, detail=str(exc)) from exc


def _resolve_db3() -> str:
    """Resolve the portfolio database (db3) path."""
    db3 = get_db3()
    if not db3:
        raise HTTPException(
            status_code=400,
            detail="Portfolio database not configured. Import transactions first."
        )
    try:
        return str(_DB_PATH_POLICY.authorize_database(db3))
    except PathPolicyError as exc:
        raise HTTPException(
            status_code=400,
            detail="Portfolio database not found. Import transactions first."
        ) from exc


def _new_backtest_id() -> str:
    return f"{datetime.now():%Y%m%d_%H%M%S}_{uuid4().hex[:8]}"


def _backtest_directory(backtest_id: str, *, require_existing: bool) -> Path:
    if not _BACKTEST_ID.fullmatch(backtest_id):
        raise HTTPException(status_code=404, detail="Backtest not found.")
    raw_candidate = _BACKTEST_ROOT / backtest_id
    if raw_candidate.is_symlink():
        raise HTTPException(status_code=404, detail="Backtest not found.")
    candidate = raw_candidate.resolve(strict=False)
    if candidate.parent != _BACKTEST_ROOT:
        raise HTTPException(status_code=404, detail="Backtest not found.")
    if require_existing and not candidate.is_dir():
        raise HTTPException(status_code=404, detail="Backtest not found.")
    return candidate


def _enforce_export_size(content: bytes) -> bytes:
    if len(content) > _APP_SETTINGS.max_export_bytes:
        raise HTTPException(413, "Generated export exceeds the configured size limit")
    return content


def _enforce_backtest_artifact_size(content: bytes) -> bytes:
    if len(content) > _APP_SETTINGS.max_backtest_artifact_bytes:
        raise HTTPException(
            413,
            "Generated backtest artifact exceeds the configured size limit",
        )
    return content


def _resolve_risk_free_rate(explicit_rf: float, base_currency: str) -> float:
    """Return the risk-free rate: explicit if > 0, else auto-detect from inflation."""
    if explicit_rf > 0:
        return explicit_rf
    try:
        db2 = get_db2()
        if db2:
            return get_risk_free_rate(db2, base_currency or "EUR")
    except Exception:
        pass
    return 0.02  # fallback


def _validate_base_currency(base_currency: str) -> str:
    """Validate and return the base currency code.

    Returns empty string on empty input (native currency).
    Raises HTTPException for invalid currency codes.
    """
    if not base_currency:
        return ""
    bc = base_currency.upper()
    valid = {c["code"] for c in get_available_display_currencies()}
    if bc not in valid:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid base currency: '{base_currency}'. "
                    f"Available currencies: {', '.join(sorted(valid))}"
        )
    return bc


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------


class AllocationSpec(BaseModel):
    mode: Literal["weight", "shares", "value"] = "weight"
    value: float


class BacktestRunRequest(BaseModel):
    db_path: str = ""
    portfolio: dict[str, AllocationSpec]
    start_date: str = Field(..., description="YYYY-MM-DD")
    end_date: str = Field(..., description="YYYY-MM-DD")
    benchmark_ticker: str = ""
    benchmark_mode: Literal["ticker", "portfolio"] = "ticker"
    base_currency: str = Field(default="", description="Target currency for returns (e.g. EUR, USD). Empty = native.")
    initial_capital: float = 0.0
    risk_free_rate: float = 0.0


class CSVBacktestRequest(BaseModel):
    db_path: str = ""
    csv_content: str = Field(..., description="Raw CSV string")
    benchmark_ticker: str = ""
    benchmark_mode: Literal["ticker", "portfolio"] = "ticker"
    base_currency: str = Field(default="", description="Target currency for returns (e.g. EUR, USD). Empty = native.")
    durations: list[str] = ["1yr", "2yr", "3yr", "5yr", "10yr"]
    initial_capital: float = 0.0
    risk_free_rate: float = 0.0


class RollingScreeningRequest(BaseModel):
    db_path: str = ""
    criteria: list[dict]
    columns: list[str]
    computed_columns: list[dict] = []
    cadence: str = "monthly"
    durations: list[str] = ["1yr", "2yr", "3yr", "5yr", "10yr"]
    weighting_modes: list[str] = ["equal"]
    max_companies: int = 25
    ranking_algorithm: str = "none"
    ranking_rules: list[dict] = []
    benchmark_ticker: str = ""
    benchmark_mode: Literal["ticker", "portfolio"] = "ticker"
    base_currency: str = Field(default="", description="Target currency for returns (e.g. EUR, USD). Empty = native.")
    initial_capital: float = 0.0
    risk_free_rate: float = 0.0
    start_period: str | None = None
    end_period: str | None = None


class RollingExportRequest(BaseModel):
    """Request body for exporting rolling backtest results to XLSX."""
    rolling_result: dict[str, Any]


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------


@router.get("/db-path")
def get_db_path() -> dict:
    """Return a stable identifier for the default database."""
    try:
        _resolve_db()
        return {"db_path": "default"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/available-tickers")
def get_available_tickers(
    db_path: str = Query(default="", description="Database path"),
) -> dict:
    """Return distinct tickers for autocomplete in the portfolio builder.

    Queries ``CompanyInfo.Company_Ticker`` for speed (smaller table).
    """
    try:
        resolved = _resolve_db(db_path)
        conn = connect_read(resolved)
        try:
            rows = conn.execute(
                "SELECT DISTINCT Company_Ticker FROM CompanyInfo "
                "WHERE Company_Ticker IS NOT NULL AND Company_Ticker != '' "
                "ORDER BY Company_Ticker"
            ).fetchall()
            return {"tickers": [r[0] for r in rows]}
        finally:
            conn.close()
    except HTTPException:
        raise
    except Exception as e:
        logger.error("available-tickers failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/base-currencies")
def get_base_currencies() -> dict:
    """Return available currencies for base currency selection."""
    currencies = get_available_display_currencies()
    return {"currencies": currencies}


@router.get("/list")
def list_backtests() -> dict:
    """List saved backtest results."""
    if not _BACKTEST_ROOT.exists():
        return {"backtests": []}
    items = []
    for d in sorted(_BACKTEST_ROOT.iterdir(), reverse=True):
        if _BACKTEST_ID.fullmatch(d.name) and d.is_dir() and not d.is_symlink():
            zip_file = d / "backtest.zip"
            items.append({
                "id": d.name,
                "path": d.name,
                "created": d.name,
                "has_zip": zip_file.exists(),
            })
    return {"backtests": items}


@router.get("/download/{backtest_id}")
def download_backtest(backtest_id: str):
    """Serve a previously saved backtest ZIP file."""
    zip_path = _backtest_directory(
        backtest_id,
        require_existing=True,
    ) / "backtest.zip"
    if not zip_path.is_file() or zip_path.is_symlink():
        raise HTTPException(status_code=404, detail="Backtest not found.")
    return FileResponse(
        str(zip_path),
        media_type="application/zip",
        filename=f"backtest_{backtest_id}.zip",
    )


@router.post("/run")
async def run_backtest(request: BacktestRunRequest = Body(...)) -> dict:
    """Run a single backtest.  Results are saved to disk; only a summary
    and the download path are returned to the client."""
    db = _resolve_db(request.db_path)

    portfolio: dict[str, dict] = {
        tk: {"mode": spec.mode, "value": spec.value}
        for tk, spec in request.portfolio.items()
    }

    base_currency = _validate_base_currency(request.base_currency)

    db3 = ""
    if request.benchmark_mode == "portfolio":
        db3 = _resolve_db3()
        if not base_currency:
            base_currency = "EUR"

    async with _semaphore:
        try:
            result = await asyncio.wait_for(
                asyncio.to_thread(
                    _bt.run_backtest_web,
                    db_path=db,
                    portfolio=portfolio,
                    start_date=request.start_date,
                    end_date=request.end_date,
                    benchmark_ticker=request.benchmark_ticker,
                    benchmark_mode=request.benchmark_mode,
                    base_currency=base_currency,
                    db3_path=db3,
                    initial_capital=request.initial_capital,
                    risk_free_rate=_resolve_risk_free_rate(request.risk_free_rate, base_currency),
                ),
                timeout=120,
            )
        except asyncio.TimeoutError:
            raise HTTPException(status_code=504, detail="Backtest timed out after 120 seconds.")
        except HTTPException:
            raise
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            import traceback
            logger.error("Backtest run failed: %s\n%s", e, traceback.format_exc())
            raise HTTPException(status_code=500, detail=str(e))

    # Save result JSON to disk immediately (fast), build ZIP in background
    ts = _new_backtest_id()
    out_dir = _backtest_directory(ts, require_existing=False)

    # Save full result as JSON (daily data included)
    def _save_and_zip():
        import csv as _csv
        import io as _io
        import json as _json
        result_bytes = _json.dumps(result, default=str).encode("utf-8")
        _enforce_backtest_artifact_size(result_bytes)
        zip_bytes = build_single_backtest_zip(result)
        _enforce_backtest_artifact_size(zip_bytes)
        daily = result.get("daily")
        daily_bytes = b""
        if daily and len(daily) > 0:
            keys = list(daily[0].keys())
            stream = _io.StringIO(newline="")
            writer = _csv.DictWriter(
                stream,
                fieldnames=keys,
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(daily)
            daily_bytes = stream.getvalue().encode("utf-8")
            _enforce_backtest_artifact_size(daily_bytes)
        if sum(map(len, (result_bytes, zip_bytes, daily_bytes))) > (
            _APP_SETTINGS.max_backtest_artifact_bytes
        ):
            raise HTTPException(
                413,
                "Generated backtest files exceed the configured size limit",
            )
        out_dir.mkdir(parents=True, exist_ok=False)
        (out_dir / "result.json").write_bytes(result_bytes)
        (out_dir / "backtest.zip").write_bytes(zip_bytes)
        if daily_bytes:
            (out_dir / "per_company_per_day.csv").write_bytes(daily_bytes)

    await asyncio.to_thread(_save_and_zip)

    return {
        "id": ts,
        "status": "complete",
        "path": ts,
        "summary": build_summary(result),
        "chart_data": result.get("chart_data", {}),
        "per_company": result.get("per_company", []),
        "yearly_returns": result.get("yearly_returns", []),
        "dividends_by_year": result.get("dividends_by_year", []),
    }


@router.post("/run-from-csv")
async def run_from_csv(request: CSVBacktestRequest = Body(...)) -> dict:
    """Run a backtest set from an uploaded CSV content string."""
    db = _resolve_db(request.db_path)

    if not request.csv_content.strip():
        raise HTTPException(status_code=400, detail="CSV content is empty.")
    if len(request.csv_content.encode("utf-8")) > _APP_SETTINGS.max_upload_bytes:
        raise HTTPException(413, "CSV content exceeds the configured size limit")

    base_currency = _validate_base_currency(request.base_currency)
    db3 = ""
    if request.benchmark_mode == "portfolio":
        db3 = _resolve_db3()
        if not base_currency:
            base_currency = "EUR"

    async with _semaphore:
        try:
            result = await asyncio.wait_for(
                asyncio.to_thread(
                    _bt.run_backtest_set_web,
                    db_path=db,
                    csv_content=request.csv_content,
                    durations=request.durations,
                    benchmark_ticker=request.benchmark_ticker,
                    benchmark_mode=request.benchmark_mode,
                    base_currency=base_currency,
                    db3_path=db3,
                    initial_capital=request.initial_capital,
                    risk_free_rate=_resolve_risk_free_rate(request.risk_free_rate, base_currency),
                ),
                timeout=120,
            )
        except asyncio.TimeoutError:
            raise HTTPException(status_code=504, detail="Backtest set timed out after 120 seconds.")
        except HTTPException:
            raise
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        except Exception as e:
            logger.error("CSV backtest set failed: %s", e)
            raise HTTPException(status_code=500, detail=str(e))

    # Save results to disk as JSON + summary
    import io as _io
    ts = _new_backtest_id()
    out_dir = _backtest_directory(ts, require_existing=False)

    result_bytes = json.dumps(result, default=str).encode("utf-8")
    _enforce_backtest_artifact_size(result_bytes)

    # Build a simple ZIP with summary
    zip_buf = _io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("result.json", json.dumps(result, default=str))
        agg = result.get("aggregate", {})
        zf.writestr("summary.json", json.dumps({
            "total_runs": agg.get("total_runs", 0),
            "successful": agg.get("successful", 0),
            "failed": agg.get("failed", 0),
        }, indent=2))
    zip_bytes = _enforce_backtest_artifact_size(zip_buf.getvalue())
    if (
        len(result_bytes) + len(zip_bytes)
        > _APP_SETTINGS.max_backtest_artifact_bytes
    ):
        raise HTTPException(
            413,
            "Generated backtest files exceed the configured size limit",
        )
    out_dir.mkdir(parents=True, exist_ok=False)
    (out_dir / "result.json").write_bytes(result_bytes)
    (out_dir / "backtest.zip").write_bytes(zip_bytes)

    return {
        "id": ts,
        "status": "complete",
        "path": ts,
        "aggregate": result.get("aggregate", {}),
    }


@router.get("/rolling-periods")
def get_rolling_periods(
    db_path: str = Query(default="", description="Database path"),
    cadence: str = Query(default="monthly", description="monthly|quarterly|yearly"),
    start_period: str | None = Query(default=None, description="YYYY-MM"),
    end_period: str | None = Query(default=None, description="YYYY-MM"),
    durations: str = Query(default="1yr,2yr,3yr,5yr,10yr"),
    weighting_modes: str = Query(default="equal", description="comma-separated"),
    financial_statements_table: str = Query(default="FinancialStatements"),
) -> dict:
    """Return available screening periods and estimated backtest count."""
    db = _resolve_db(db_path)
    try:
        periods = _bt._discover_screening_periods(
            db, cadence, start_period, end_period,
            financial_statements_table=financial_statements_table,
        )
        dur_list = [d.strip() for d in durations.split(",") if d.strip()]
        wm_list = [w.strip() for w in weighting_modes.split(",") if w.strip()]
        estimated_backtests = len(periods) * len(dur_list) * len(wm_list)
        return {
            "periods": periods,
            "count": len(periods),
            "estimated_backtests": estimated_backtests,
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("rolling-periods failed: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/run-rolling")
async def run_rolling(
    request: RollingScreeningRequest,
    http_request: Request,
) -> StreamingResponse:
    """Run a rolling screening backtest with SSE progress streaming."""
    db = _resolve_db(request.db_path)

    base_currency = _validate_base_currency(request.base_currency)
    db3 = ""
    if request.benchmark_mode == "portfolio":
        db3 = _resolve_db3()
        if not base_currency:
            base_currency = "EUR"

    progress_queue: queue.Queue = queue.Queue()
    cancel_event = threading.Event()

    async def event_generator():
        loop = asyncio.get_event_loop()
        task = asyncio.ensure_future(
            asyncio.to_thread(
                _bt.run_screening_backtest_rolling,
                db_path=db,
                criteria=request.criteria,
                columns=request.columns,
                cadence=request.cadence,
                durations=request.durations,
                weighting_modes=request.weighting_modes,
                max_companies=request.max_companies,
                ranking_algorithm=request.ranking_algorithm,
                ranking_rules=request.ranking_rules,
                computed_columns=request.computed_columns,
                benchmark_ticker=request.benchmark_ticker,
                benchmark_mode=request.benchmark_mode,
                base_currency=base_currency,
                db3_path=db3,
                initial_capital=request.initial_capital,
                risk_free_rate=_resolve_risk_free_rate(request.risk_free_rate, base_currency),
                start_period=request.start_period,
                end_period=request.end_period,
                progress_queue=progress_queue,
                cancel_event=cancel_event,
            )
        )
        final_result = None
        try:
            while True:
                if await http_request.is_disconnected():
                    cancel_event.set()
                    break

                try:
                    msg = await loop.run_in_executor(
                        None, progress_queue.get, True, 1.0,
                    )
                except queue.Empty:
                    if task.done():
                        break
                    continue

                # Don't stream the full result to the client — too large
                if msg.get("type") == "result":
                    final_result = msg
                    break
                if msg.get("type") == "error":
                    yield f"data: {json.dumps(msg)}\n\n"
                    break

                yield f"data: {json.dumps(msg)}\n\n"
        finally:
            cancel_event.set()
            if not task.done():
                task.cancel()

        if task.done() and not task.cancelled():
            exc = task.exception()
            if exc is not None:
                error_msg = str(exc)
                if "cancelled" in error_msg.lower():
                    yield f"data: {json.dumps({'type': 'error', 'message': 'cancelled'})}\n\n"
                else:
                    logger.error("Rolling backtest failed: %s", error_msg)
                    yield f"data: {json.dumps({'type': 'error', 'message': 'Backtest failed'})}\n\n"

        # Build ZIP and save to disk, then yield download info
        if final_result is not None:
            try:
                saved_path = await loop.run_in_executor(
                    None,
                    save_rolling_backtest_zip,
                    final_result,
                    str(_BACKTEST_ROOT),
                    _APP_SETTINGS.max_backtest_artifact_bytes,
                )
                backtest_id = saved_path.split("/")[-1] if "/" in saved_path else saved_path.split("\\")[-1]
                agg = final_result.get("aggregate", {})
                cfg = final_result.get("config", {})
                yield f"data: {json.dumps({'type': 'result', 'id': backtest_id, 'path': backtest_id, 'aggregate': agg, 'config': cfg})}\n\n"
            except ExportSizeLimitExceeded as exc:
                limit_mib = exc.limit_bytes // (1024 * 1024)
                logger.warning(
                    "Rolling backtest archive exceeded limit: attempted=%d limit=%d",
                    exc.attempted_bytes,
                    exc.limit_bytes,
                )
                message = (
                    "Backtest completed, but its download archive exceeded "
                    f"the {limit_mib} MiB limit. Reduce the period range or "
                    "increase EDINET_MAX_BACKTEST_ARTIFACT_BYTES and restart."
                )
                yield f"data: {json.dumps({'type': 'error', 'message': message})}\n\n"
            except Exception as e:
                logger.error("Failed to save rolling backtest ZIP: %s", e, exc_info=True)
                yield f"data: {json.dumps({'type': 'error', 'message': 'Failed to save results'})}\n\n"

    async with _semaphore:
        return StreamingResponse(
            event_generator(),
            media_type="text/event-stream",
        )


@router.post("/export-rolling-xlsx")
async def export_rolling_xlsx(request: RollingExportRequest) -> StreamingResponse:
    """Export rolling backtest results as a multi-sheet XLSX workbook.

    Sheet 1: "Summary" — aggregate statistics, by-weighting, benchmark.
    Subsequent sheets: one per screening period, containing per-backtest
    data for each duration × weighting combination.
    """
    import io
    from datetime import datetime

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    request_bytes = json.dumps(request.rolling_result, default=str).encode("utf-8")
    _enforce_export_size(request_bytes)
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────────
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E2F2", end_color="D9E2F2", fill_type="solid")
    pct_fmt = '0.00%'
    num_fmt = '0.000'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    def auto_width(ws, min_w=10, max_w=40):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))

    result = request.rolling_result
    agg = result.get("aggregate", {})
    config = result.get("config", {})
    results_list = result.get("results", [])

    # ── Summary sheet ─────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Config info
    ws_sum.cell(row=1, column=1, value="Rolling Screening Backtest").font = Font(bold=True, size=14)
    ws_sum.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    row = 4
    for key, label in [("cadence", "Cadence"), ("durations", "Durations"),
                       ("weighting_modes", "Weighting"), ("max_companies", "Max Companies"),
                       ("benchmark_ticker", "Benchmark"), ("start_period", "Start"),
                       ("end_period", "End")]:
        ws_sum.cell(row=row, column=1, value=label).font = Font(bold=True)
        val = config.get(key, "")
        ws_sum.cell(row=row, column=2, value=str(val) if not isinstance(val, list) else ", ".join(val))
        row += 1

    row += 1
    ws_sum.cell(row=row, column=1, value=f"Total Backtests: {agg.get('total_runs', 0)}").font = Font(bold=True)
    ws_sum.cell(row=row + 1, column=1, value=f"Successful: {agg.get('successful', 0)}")
    ws_sum.cell(row=row + 2, column=1, value=f"Failed: {agg.get('failed', 0)}")
    ws_sum.cell(row=row + 3, column=1, value=f"Periods: {agg.get('periods', 0)}")

    # Summary by duration/weighting
    row += 5
    by_w = agg.get("by_weighting", {})
    if by_w:
        ws_sum.cell(row=row, column=1, value="Returns by Duration & Weighting").font = Font(bold=True, size=12)
        row += 1
        headers = ["Weighting", "Duration", "Count", "Mean Ann. Return", "Median Ann. Return", "Mean Sharpe"]
        for c, h in enumerate(headers, 1):
            ws_sum.cell(row=row, column=c, value=h)
        style_header(ws_sum, row, len(headers))
        row += 1
        for wm, dur_data in by_w.items():
            for dur, stats in dur_data.items():
                ws_sum.cell(row=row, column=1, value=wm)
                ws_sum.cell(row=row, column=2, value=dur)
                ws_sum.cell(row=row, column=3, value=stats.get("count", 0))
                ws_sum.cell(row=row, column=4, value=stats.get("mean_return", 0))
                ws_sum.cell(row=row, column=5, value=stats.get("median_return", 0))
                ws_sum.cell(row=row, column=6, value=stats.get("mean_sharpe", 0))
                for c in range(1, 7):
                    ws_sum.cell(row=row, column=c).border = thin_border
                ws_sum.cell(row=row, column=4).number_format = pct_fmt
                ws_sum.cell(row=row, column=5).number_format = pct_fmt
                ws_sum.cell(row=row, column=6).number_format = num_fmt
                row += 1

        # Benchmark comparison
        bc = agg.get("benchmark_comparison")
        if bc:
            row += 1
            ws_sum.cell(row=row, column=1, value="Benchmark Comparison").font = Font(bold=True, size=12)
            row += 1
            ws_sum.cell(row=row, column=1, value=f"Outperformed: {bc.get('outperformed', 0)}")
            ws_sum.cell(row=row + 1, column=1, value=f"Underperformed: {bc.get('underperformed', 0)}")
            ws_sum.cell(row=row + 2, column=1, value=f"Win Rate: {bc.get('win_rate', 0) * 100:.1f}%")

        auto_width(ws_sum)

    # ── Per-period sheets ──────────────────────────────────────────
    for period_result in results_list:
        period = period_result.get("period", "unknown")
        # Sheet names max 31 chars
        sheet_name = period[:10].replace("-", "")[:31]
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name[:28] + f"_{len(wb.sheetnames)}"
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(row=1, column=1, value=f"Period: {period}").font = Font(bold=True, size=13)
        ws.cell(row=2, column=1,
                value=f"Tickers ({period_result.get('ticker_count', 0)}): " +
                      ", ".join(period_result.get("tickers", [])))

        backtests = period_result.get("backtests", {})
        for wm, dur_data in backtests.items():
            row = 4
            ws.cell(row=row, column=1, value=f"Weighting: {wm}").font = Font(bold=True, size=12)
            row += 1

            headers = ["Duration", "Total Return", "Ann. Return", "Sharpe",
                       "Max Drawdown", "Benchmark Return", "Excess Return",
                       "Volatility", "Info Ratio", "Start", "End"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=row, column=c, value=h)
            style_header(ws, row, len(headers))
            row += 1

            for dur, bt in dur_data.items():
                m = bt.get("metrics") if isinstance(bt, dict) else None
                if m is None:
                    ws.cell(row=row, column=1, value=dur)
                    ws.cell(row=row, column=2, value="Failed")
                    row += 1
                    continue

                ws.cell(row=row, column=1, value=dur)
                ws.cell(row=row, column=2, value=m.get("total_return", 0))
                ws.cell(row=row, column=3, value=m.get("annualized_return", 0))
                ws.cell(row=row, column=4, value=m.get("sharpe_ratio", 0))
                ws.cell(row=row, column=5, value=m.get("max_drawdown", 0))
                bench = m.get("benchmark_total_return")
                ws.cell(row=row, column=6, value=bench if bench is not None else "N/A")
                if bench is not None and m.get("total_return") is not None:
                    ws.cell(row=row, column=7, value=m["total_return"] - bench)
                else:
                    ws.cell(row=row, column=7, value="N/A")
                ws.cell(row=row, column=8, value=m.get("volatility", 0))
                info = m.get("information_ratio")
                ws.cell(row=row, column=9, value=info if info is not None else "N/A")
                ws.cell(row=row, column=10, value=m.get("start_date", ""))
                ws.cell(row=row, column=11, value=m.get("end_date", ""))

                for c in range(1, 12):
                    ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=2).number_format = pct_fmt
                ws.cell(row=row, column=3).number_format = pct_fmt
                ws.cell(row=row, column=4).number_format = num_fmt
                ws.cell(row=row, column=5).number_format = pct_fmt
                if bench is not None:
                    ws.cell(row=row, column=6).number_format = pct_fmt
                    ws.cell(row=row, column=7).number_format = pct_fmt
                ws.cell(row=row, column=8).number_format = pct_fmt
                if info is not None:
                    ws.cell(row=row, column=9).number_format = num_fmt
                row += 1

            # Per-company breakdown if available
            for dur, bt in dur_data.items():
                per_co = bt.get("per_company") if isinstance(bt, dict) else None
                if per_co and len(per_co) > 0:
                    row += 1
                    ws.cell(row=row, column=1,
                            value=f"{dur} — Per-Company Breakdown").font = Font(bold=True)
                    row += 1
                    co_headers = ["Ticker", "Total Return", "Price Return",
                                  "Dividend Return", "Weight",
                                  "Wtd Price", "Wtd Div", "Wtd Total",
                                  "Start Price", "End Price"]
                    # Add cash columns if available
                    has_capital = any(
                        co.get("capital_invested") is not None for co in per_co
                    )
                    if has_capital:
                        co_headers += ["Capital", "Shares", "Divs Received", "Market Value"]
                    for c, h in enumerate(co_headers, 1):
                        ws.cell(row=row, column=c, value=h)
                    style_header(ws, row, len(co_headers))
                    row += 1
                    for co in per_co:
                        col = 1
                        ws.cell(row=row, column=col, value=co.get("Ticker", "")); col += 1
                        ws.cell(row=row, column=col, value=co.get("total_return", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("price_return", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("dividend_return", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("weight", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("weighted_price", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("weighted_dividend", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("weighted_total", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("start_price", 0)); col += 1
                        ws.cell(row=row, column=col, value=co.get("end_price", 0)); col += 1
                        if has_capital:
                            ws.cell(row=row, column=col, value=co.get("capital_invested", 0)); col += 1
                            ws.cell(row=row, column=col, value=co.get("shares_purchased", 0)); col += 1
                            ws.cell(row=row, column=col, value=co.get("dividends_received", 0)); col += 1
                            ws.cell(row=row, column=col, value=co.get("market_value", 0)); col += 1
                        for c in range(1, col):
                            ws.cell(row=row, column=c).border = thin_border
                        for c in [2, 3, 4, 5, 6, 7, 8]:
                            ws.cell(row=row, column=c).number_format = pct_fmt
                        row += 1
                    break  # Only show per-company for first duration

        auto_width(ws)

    # ── Write to bytes ───────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    content = _enforce_export_size(output.getvalue())

    filename = f"rolling_backtest_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export-rolling-zip")
async def export_rolling_zip(request: RollingExportRequest) -> StreamingResponse:
    """Export rolling backtest results as a ZIP archive.

    The ZIP contains:

    - ``summary.txt`` / ``summary.csv`` — aggregate statistics
    - ``heatmap.csv`` — returns matrix (periods × durations × weightings)
    - ``backtests/{period}_{weighting}_{duration}/`` — per‑backtest files:
      ``report.txt``, chart PNGs, ``per_company.csv``,
      ``yearly_returns.csv``, ``dividends.csv``
    """
    import io as _io
    from datetime import datetime as _datetime

    request_bytes = json.dumps(request.rolling_result, default=str).encode("utf-8")
    _enforce_export_size(request_bytes)
    zip_bytes = await asyncio.to_thread(
        build_rolling_zip, request.rolling_result,
    )
    _enforce_export_size(zip_bytes)

    filename = f"rolling_backtest_{_datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    return StreamingResponse(
        _io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
